"""Utilities for generating the KodeKloud usage report."""

from typing import Any, Dict, List, Tuple

import pandas as pd
import json
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from ..external_services import blob
from ..core import settings

warnings.simplefilter("ignore")


def convert_to_hours(value: Any) -> float:
    """Convert textual hour/minute values into a float representing hours."""

    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value = str(value).strip().lower()
    if "hour" in value:
        num = re.findall(r"[\d\.]+", value)
        return float(num[0]) if num else 0.0
    if "minute" in value:
        num = re.findall(r"[\d\.]+", value)
        return float(num[0]) / 60 if num else 0.0
    return 0.0


def load_input_files(admin_path: str, activity_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Load the raw Excel input files and normalise basic columns."""

    admin_df = pd.read_excel(admin_path)
    activity_df = pd.read_excel(activity_path)

    admin_df["Email"] = admin_df["Email"].str.strip().str.lower()
    activity_df["Email"] = activity_df["Email"].str.strip().str.lower()

    admin_df = admin_df[["Name", "Email", "Program", "License Accepted"]]
    activity_df = activity_df[["Email", "Lessons Completed", "Video Hours Watched", "Labs Completed"]]

    # remove LPC users
    admin_df = admin_df[admin_df["Program"].str.strip().str.upper() != "LPC"]
    return admin_df, activity_df


def merge_activity_data(admin_df: pd.DataFrame, activity_df: pd.DataFrame) -> pd.DataFrame:
    """Merge admin and activity dataframes and derive status fields."""

    merged = pd.merge(admin_df, activity_df, on="Email", how="left")

    merged["Lessons Completed"] = merged["Lessons Completed"].fillna(0).astype(int)
    merged["Video Hours Watched"] = merged["Video Hours Watched"].apply(convert_to_hours)
    merged["Labs Completed"] = merged["Labs Completed"].fillna(0).astype(int)

    def activity_status(row: pd.Series) -> str:
        if (
            row["Lessons Completed"] == 0
            and row["Video Hours Watched"] == 0
            and row["Labs Completed"] == 0
        ):
            return "No activity or progress"
        return ""

    merged["Status"] = merged.apply(activity_status, axis=1)

    def license_display(val: Any) -> str:
        if isinstance(val, str) and val.strip().lower() == "no":
            return "X"
        return "\u2713"

    merged["License Accepted Display"] = merged["License Accepted"].apply(license_display)

    columns = [
        "Name",
        "Email",
        "Program",
        "Lessons Completed",
        "Video Hours Watched",
        "Labs Completed",
        "License Accepted Display",
        "Status",
    ]
    display_df = merged[columns].rename(columns={"License Accepted Display": "License Accepted"})
    return display_df


def format_excel(df: pd.DataFrame, output_excel_path: str) -> None:
    """Save the DataFrame to Excel and apply conditional formatting."""

    df.to_excel(output_excel_path, index=False)

    wb = load_workbook(output_excel_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        status = row[7].value
        license_col = row[6].value
        if status == "No activity or progress":
            for cell in row:
                cell.fill = red_fill
        elif license_col == "X":
            for cell in row:
                cell.fill = orange_fill

    wb.save(output_excel_path)


def write_json(df: pd.DataFrame, output_json_path: str) -> List[Dict[str, Any]]:
    """Write the DataFrame to a JSON file and return the data."""

    json_data = df.to_dict(orient="records")
    with open(output_json_path, "w") as f:
        json.dump(json_data, f, indent=2)
    return json_data


def upload_json(local_json_path: str) -> None:
    """Upload the JSON file to Azure Blob Storage."""

    blob.upload_file_to_blob(
        settings.CONTAINER_INPUTS,
        settings.JSON_BLOB_PATH,
        local_json_path,
    )


def generate_report(
    admin_path: str, activity_path: str, output_excel_path: str, output_json_path: str
) -> List[Dict[str, Any]]:
    """Generate the usage report in Excel and JSON formats."""

    admin_df, activity_df = load_input_files(admin_path, activity_path)
    report_df = merge_activity_data(admin_df, activity_df)
    format_excel(report_df, output_excel_path)
    json_data = write_json(report_df, output_json_path)
    upload_json(output_json_path)
    return json_data
